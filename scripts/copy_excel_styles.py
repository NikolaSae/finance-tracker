import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

def rgb_to_color(rgb):
    # Ensure the RGB value is in the correct format (e.g., "RRGGBB" or "AARRGGBB")
    if rgb and isinstance(rgb, str):
        if rgb.startswith('FF'):  # Strip the alpha channel (FF) if it exists
            return Color(rgb=rgb[2:])
        return Color(rgb=rgb)
    return None

def copy_excel_styles(input_file, output_file):
    # Load the original workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active  # First sheet

    # Create a new workbook
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Copy all cells, values, formats, and merged cells
    for row in ws.iter_rows():
        for cell in row:
            # Copy the value of the cell
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy font style
            if cell.font:
                new_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    underline=cell.font.underline,
                    color=cell.font.color
                )

            # Copy fill (background color)
            if cell.fill:
                start_color = rgb_to_color(cell.fill.start_color.rgb) if cell.fill.start_color else None
                end_color = rgb_to_color(cell.fill.end_color.rgb) if cell.fill.end_color else None
                new_cell.fill = PatternFill(
                    start_color=start_color,
                    end_color=end_color,
                    fill_type=cell.fill.fill_type
                )

            # Copy border
            if cell.border:
                new_cell.border = Border(
                    left=Side(style=cell.border.left.style, color=rgb_to_color(cell.border.left.color.rgb) if cell.border.left.color else None),
                    right=Side(style=cell.border.right.style, color=rgb_to_color(cell.border.right.color.rgb) if cell.border.right.color else None),
                    top=Side(style=cell.border.top.style, color=rgb_to_color(cell.border.top.color.rgb) if cell.border.top.color else None),
                    bottom=Side(style=cell.border.bottom.style, color=rgb_to_color(cell.border.bottom.color.rgb) if cell.border.bottom.color else None)
                )

            # Copy alignment
            if cell.alignment:
                new_cell.alignment = openpyxl.styles.Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )

            # Copy merged cells
            if cell.coordinate in ws.merged_cells:
                # This cell is part of a merged block, so we copy it to the new sheet
                merged_cells = [merged for merged in ws.merged_cells if cell.coordinate in merged]
                for merged in merged_cells:
                    new_ws.merge_cells(str(merged))

    # Save the new workbook with all the data and styles
    new_wb.save(output_file)

# Pozivanje funkcije sa ulaznim i izlaznim fajlom
input_file = "E:/xampp-8-telekom/htdocs/finance-tracker/scripts/data/yourfile.xlsx"
output_file = 'data/output_with_styles.xlsx'  # Put za izlazni fajl
copy_excel_styles(input_file, output_file)
