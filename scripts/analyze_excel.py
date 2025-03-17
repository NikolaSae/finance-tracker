import openpyxl
from openpyxl.styles import Font, Border, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# Function to extract formatting details for all filled cells
def analyze_excel(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Kreiramo novi radni list za rezultate
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append([
        "Row", "Column", "Value", "Font Name", "Font Size", "Font Color", 
        "Bold", "Italic", "Underline", "Border Left", "Border Right", 
        "Border Top", "Border Bottom", "Fill Color", "Merged Cells"
    ])

    # Prolazimo kroz sve ćelije na radnom listu
    for row in ws.iter_rows():
        for cell in row:
            # Proveravamo ako je ćelija deo spojenog bloka
            merge_info = ""
            if cell.coordinate in ws.merged_cells:
                merge_info = "Yes"
            else:
                merge_info = "No"
            
            # Dodajemo podatke o formatu i vrednosti
            new_ws.append([
                str(cell.row), str(cell.column), str(cell.value or ""),
                str(cell.font.name), str(cell.font.size), str(cell.font.color or ""),
                str(cell.font.bold), str(cell.font.italic), str(cell.font.underline),
                str(cell.border.left.style), str(cell.border.right.style),
                str(cell.border.top.style), str(cell.border.bottom.style),
                str(cell.fill.start_color.rgb if cell.fill.start_color else ""),
                merge_info
            ])

    # Snimanje rezultata u novi Excel fajl
    new_wb.save(output_file)
    print(f"Analysis saved to {output_file}")


# Example usage
input_file = "E:/xampp-8-telekom/htdocs/finance-tracker/scripts/data/yourfile.xlsx"
  # Replace with your actual file path
output_file = "E:/xampp-8-telekom/htdocs/finance-tracker/scripts/data/formatting_analysis.xlsx"
analyze_excel(input_file, output_file)
